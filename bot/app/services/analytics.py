from __future__ import annotations

from datetime import datetime
import logging

from aiogram.types import User
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.config import BASE_DIR


ANALYTICS_PATH = BASE_DIR / "data" / "analytics.xlsx"

HEADERS = [
    "Telegram ID",
    "Username",
    "Имя в Telegram",
    "Первый вход",
    "Последняя активность",
    "Источник",
    "Кол-во входов",
    "Забрал материалы",
    "Мини-тест начат",
    "Мини-тест завершен",
    "Мини-тест брошен",
    "Тест с разбором начат",
    "Тест с разбором завершен",
    "Тест с разбором брошен",
    "Последнее действие",
    "Имя из формы",
    "Возраст / для кого",
    "Цель",
    "Уровень",
    "Боль",
    "Формат",
    "Контакт",
    "Комментарий",
]

COLUMN = {name: index + 1 for index, name in enumerate(HEADERS)}

EVENT_LABELS = {
    "start": "Зашел в бота",
    "materials_taken": "Забрал материалы",
    "mini_test_started": "Начал мини-тест",
    "mini_test_completed": "Прошел мини-тест",
    "mini_test_abandoned": "Вышел из мини-теста",
    "diagnostic_started": "Начал тест с разбором",
    "diagnostic_completed": "Прошел тест с разбором",
    "diagnostic_abandoned": "Вышел из теста с разбором",
}

EVENT_COLUMNS = {
    "materials_taken": "Забрал материалы",
    "mini_test_started": "Мини-тест начат",
    "mini_test_completed": "Мини-тест завершен",
    "mini_test_abandoned": "Мини-тест брошен",
    "diagnostic_started": "Тест с разбором начат",
    "diagnostic_completed": "Тест с разбором завершен",
    "diagnostic_abandoned": "Тест с разбором брошен",
}

FORM_FIELDS = {
    "name": "Имя из формы",
    "age_target": "Возраст / для кого",
    "goal": "Цель",
    "level": "Уровень",
    "pain": "Боль",
    "format": "Формат",
    "contact": "Контакт",
    "comment": "Комментарий",
}


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _user_id(user: User | None) -> str | None:
    if not user:
        return None
    return str(user.id)


def _username(user: User | None) -> str:
    if not user or not user.username:
        return ""
    return f"@{user.username}"


def _full_name(user: User | None) -> str:
    if not user:
        return ""
    return user.full_name


def _prepare_sheet(sheet: Worksheet) -> None:
    sheet.title = "Пользователи"
    sheet.append(HEADERS)
    sheet.freeze_panes = "A2"

    widths = {
        "A": 14,
        "B": 18,
        "C": 24,
        "D": 20,
        "E": 20,
        "F": 18,
        "G": 14,
        "H": 18,
        "I": 18,
        "J": 20,
        "K": 18,
        "L": 22,
        "M": 24,
        "N": 22,
        "O": 26,
        "P": 22,
        "Q": 24,
        "R": 26,
        "S": 20,
        "T": 24,
        "U": 24,
        "V": 26,
        "W": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _load_or_create_workbook() -> Workbook:
    ANALYTICS_PATH.parent.mkdir(parents=True, exist_ok=True)

    if ANALYTICS_PATH.exists():
        workbook = load_workbook(ANALYTICS_PATH)
        sheet = workbook.active
        existing_headers = [sheet.cell(row=1, column=col).value for col in range(1, len(HEADERS) + 1)]
        if existing_headers != HEADERS:
            for index, header in enumerate(HEADERS, start=1):
                sheet.cell(row=1, column=index, value=header)
        return workbook

    workbook = Workbook()
    _prepare_sheet(workbook.active)
    return workbook


def _find_user_row(sheet: Worksheet, telegram_id: str) -> int | None:
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=COLUMN["Telegram ID"]).value) == telegram_id:
            return row
    return None


def _get_or_create_user_row(sheet: Worksheet, user: User | None, source: str | None) -> int | None:
    telegram_id = _user_id(user)
    if telegram_id is None:
        return None

    row = _find_user_row(sheet, telegram_id)
    current_time = _now()

    if row is None:
        row = sheet.max_row + 1
        sheet.cell(row=row, column=COLUMN["Telegram ID"], value=telegram_id)
        sheet.cell(row=row, column=COLUMN["Первый вход"], value=current_time)
        sheet.cell(row=row, column=COLUMN["Кол-во входов"], value=0)

    sheet.cell(row=row, column=COLUMN["Username"], value=_username(user))
    sheet.cell(row=row, column=COLUMN["Имя в Telegram"], value=_full_name(user))
    sheet.cell(row=row, column=COLUMN["Последняя активность"], value=current_time)
    if source:
        sheet.cell(row=row, column=COLUMN["Источник"], value=source)

    return row


def _increment(sheet: Worksheet, row: int, header: str) -> None:
    cell = sheet.cell(row=row, column=COLUMN[header])
    try:
        current_value = int(cell.value or 0)
    except (TypeError, ValueError):
        current_value = 0
    cell.value = current_value + 1


def track_user_event(
    user: User | None,
    event: str,
    *,
    source: str | None = None,
    form_data: dict | None = None,
) -> None:
    try:
        workbook = _load_or_create_workbook()
        sheet = workbook.active
        row = _get_or_create_user_row(sheet, user, source or (form_data or {}).get("source"))
        if row is None:
            return

        if event == "start":
            _increment(sheet, row, "Кол-во входов")

        event_column = EVENT_COLUMNS.get(event)
        if event_column:
            sheet.cell(row=row, column=COLUMN[event_column], value="да")

        sheet.cell(
            row=row,
            column=COLUMN["Последнее действие"],
            value=EVENT_LABELS.get(event, event),
        )

        if form_data:
            for field, header in FORM_FIELDS.items():
                value = form_data.get(field)
                if value not in (None, ""):
                    sheet.cell(row=row, column=COLUMN[header], value=str(value))

        workbook.save(ANALYTICS_PATH)
    except Exception:
        logging.exception("Could not update analytics workbook")
